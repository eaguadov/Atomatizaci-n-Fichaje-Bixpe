import os
import json
import datetime
import subprocess

try:
    from openpyxl import load_workbook
except ImportError:
    print("La librería 'openpyxl' no está instalada. Ejecuta en tu consola: pip install openpyxl")
    exit(1)

USER_HOME = os.path.expanduser("~")
DEFAULT_DIR = os.path.join(USER_HOME, "Tower Consultores SL", "Proyecto RSSI - Documentos", "General")
OUTPUT_JSON = "team_holidays.json"

def find_excel_path():
    """Busca el archivo de vacaciones soportando .xlsm o .xlsx de forma dinámica para cualquier usuario"""
    candidatos = [
        os.path.join(DEFAULT_DIR, "Vacaciones RSSI.xlsm"),
        os.path.join(DEFAULT_DIR, "Vacaciones RSSI.xlsx"),
    ]
    for c in candidatos:
        if os.path.exists(c):
            return c
            
    # Búsqueda dinámica en el directorio por si cambia el nombre ligeramente
    if os.path.exists(DEFAULT_DIR):
        for f in os.listdir(DEFAULT_DIR):
            if f.lower().startswith("vacaciones") and (f.endswith(".xlsx") or f.endswith(".xlsm")):
                return os.path.join(DEFAULT_DIR, f)
    return None

def get_holiday_reason(cell):
    """Detecta si una celda es festivo/vacaciones leyendo su texto explícito.
    No se usa detección por color porque el Excel colorea los fines de semana
    de rojo como decoración visual, causando falsos positivos."""
    val = str(cell.value or '').strip().upper()
    
    # Comprobar texto explícito (V, V25, V26, V99... y festivos)
    if val == 'V' or (val.startswith('V') and val[1:].isdigit()) or val in ['FL', 'FA', 'F']:
        return val
            
    return None

def get_month_number(text):
    """Deduce el número de mes a partir de un texto como 'ene-26'."""
    text = str(text).lower().strip()
    months = {'ene': 1, 'feb': 2, 'mar': 3, 'abr': 4, 'may': 5, 'jun': 6,
              'jul': 7, 'ago': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dic': 12}
    for m_text, num in months.items():
        if text.startswith(m_text):
            return num
    return None

def get_month_from_sheet(ws, r):
    """Detecta el número de mes buscando en las filas anteriores a NOMBRE."""
    for up_row in [r-2, r-1, r-3]:
        if up_row >= 1:
            for col in range(1, 15):
                val = ws.cell(row=up_row, column=col).value
                if isinstance(val, (datetime.date, datetime.datetime)):
                    return val.month
                if val:
                    m = get_month_number(val)
                    if m:
                        return m
    return None

def main():
    excel_path = find_excel_path()
    if not excel_path:
        print(f"ERROR: No se encontró ningún archivo de vacaciones (.xlsm o .xlsx) en:\n{DEFAULT_DIR}")
        print("¿Estás conectado a la VPN / OneDrive?")
        return

    print(f"Cargando Excel desde: {excel_path}")
    try:
        wb = load_workbook(excel_path, data_only=True)
    except PermissionError:
        print("\n" + "="*56)
        print("  AVISO: El archivo Excel está abierto en tu ordenador.")
        print("  Por favor, GUARDA y CIERRA Microsoft Excel.")
        print("="*56 + "\n")
        return
    current_year = datetime.datetime.now().year
    years_to_check = [str(current_year), str(current_year + 1)]

    team_holidays = {}

    for year_str in years_to_check:
        if year_str not in wb.sheetnames:
            print(f"Pestaña {year_str} no encontrada. Saltando.")
            continue

        print(f"Procesando pestaña: {year_str}")
        ws = wb[year_str]
        month_counter = 1

        for r in range(1, ws.max_row + 1):
            cell_val = str(ws.cell(row=r, column=1).value or '').strip()
            
            # Buscamos la palabra clave que inicia la tabla de días
            if cell_val.upper() == "NOMBRE":
                detected_m = get_month_from_sheet(ws, r)
                month_num = detected_m if detected_m else month_counter
                print(f"  -> Procesando Mes {month_num} del año {year_str} (Fila de días: {r})")
                
                # Leer empleados en las filas siguientes (hasta 15 filas)
                for er in range(r + 1, r + 15):
                    emp_name = str(ws.cell(row=er, column=1).value or '').strip()
                    
                    # Si llegamos al siguiente bloque o mes, paramos este bloque
                    if emp_name.upper() == "NOMBRE" or get_month_number(emp_name):
                        break
                    
                    # Si la fila está vacía (separador), continuamos buscando en la siguiente
                    if not emp_name:
                        continue
                    
                    if emp_name not in team_holidays:
                        team_holidays[emp_name] = {}
                    
                    # Leer columnas del 1 al 31 (B=2 hasta AF=32)
                    for c in range(2, 33):
                        day_val = ws.cell(row=r, column=c).value
                        if isinstance(day_val, (int, float)) or (isinstance(day_val, str) and str(day_val).isdigit()):
                            day = int(day_val)
                            if 1 <= day <= 31:
                                try:
                                    date_obj = datetime.date(int(year_str), month_num, day)
                                    
                                    # Omitir fines de semana (Sábado=5, Domingo=6)
                                    if date_obj.weekday() >= 5:
                                        continue
                                        
                                    target_cell = ws.cell(row=er, column=c)
                                    reason = get_holiday_reason(target_cell)
                                    
                                    if reason:
                                        date_str = date_obj.strftime("%Y-%m-%d")
                                        team_holidays[emp_name][date_str] = reason
                                except ValueError:
                                    pass # Día inválido como 30 de febrero
                                    
                month_counter += 1
                if month_counter > 12:
                    break

    # Ordenar por fecha cada diccionario de empleado
    for emp in team_holidays:
        team_holidays[emp] = dict(sorted(team_holidays[emp].items()))

    # Guardar JSON en la raíz del repositorio
    repo_dir = os.path.dirname(os.path.dirname(__file__))
    json_path = os.path.join(repo_dir, OUTPUT_JSON)
    
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(team_holidays, f, indent=4, ensure_ascii=False)
    print(f"\nArchivo maestro generado en: {json_path}")

    # Ejecutar comandos de Git para subir automáticamente
    print("\nSincronizando con GitHub (Subida Automática)...")
    try:
        subprocess.run(["git", "add", OUTPUT_JSON], cwd=repo_dir, check=True)
        commit_res = subprocess.run(["git", "commit", "-m", "chore: actualizar calendario maestro desde Excel"], cwd=repo_dir, capture_output=True)
        push_res = subprocess.run(["git", "push"], cwd=repo_dir, capture_output=True)
        
        if commit_res.returncode == 0 or (push_res and push_res.returncode == 0):
            print("¡Completado! El archivo está actualizado en GitHub. Tus compañeros ya tienen los datos.")
        else:
            print("El calendario maestro ya estaba al día en GitHub.")
    except Exception as e:
        print(f"Error durante la sincronización con GitHub: {e}")

if __name__ == "__main__":
    main()
