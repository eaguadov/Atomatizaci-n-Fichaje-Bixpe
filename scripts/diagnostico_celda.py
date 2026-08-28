"""Script de diagnóstico para investigar la celda 01/06/2026 de Eusebio."""
import os, datetime
from openpyxl import load_workbook

DEFAULT_DIR = r"C:\Users\eusebio.aguado\Tower Consultores SL\Proyecto RSSI - Documentos\General"

# Buscar el Excel
for f in os.listdir(DEFAULT_DIR):
    if f.lower().startswith("vacaciones") and (f.endswith(".xlsx") or f.endswith(".xlsm")):
        excel_path = os.path.join(DEFAULT_DIR, f)
        break

print(f"Abriendo: {excel_path}")
wb = load_workbook(excel_path, data_only=True)
ws = wb["2026"]

# Buscar el bloque de Junio (mes 6, debería ser el 6º NOMBRE)
nombre_count = 0
for r in range(1, ws.max_row + 1):
    val = str(ws.cell(row=r, column=1).value or '').strip()
    if val.upper() == "NOMBRE":
        nombre_count += 1
        if nombre_count == 6:  # Junio
            print(f"\n=== BLOQUE JUNIO (Fila NOMBRE: {r}) ===")
            # Buscar la columna del día 1
            for c in range(2, 33):
                dv = ws.cell(row=r, column=c).value
                if str(dv).strip() == "1":
                    col_dia1 = c
                    print(f"Columna del día 1: {c}")
                    break
            
            # Buscar la fila de Eusebio
            for er in range(r+1, r+15):
                emp = str(ws.cell(row=er, column=1).value or '').strip()
                if emp.lower() == "eusebio":
                    cell = ws.cell(row=er, column=col_dia1)
                    print(f"\n--- Celda de Eusebio, día 1 de Junio (Fila {er}, Col {col_dia1}) ---")
                    print(f"  Valor:      '{cell.value}'")
                    print(f"  Fill type:  '{getattr(cell.fill, 'fill_type', 'N/A')}'")
                    print(f"  Fill color index: '{getattr(cell.fill.start_color, 'index', 'N/A')}'")
                    print(f"  Fill color rgb:   '{getattr(cell.fill.start_color, 'rgb', 'N/A')}'")
                    print(f"  Fill color type:  '{getattr(cell.fill.start_color, 'type', 'N/A')}'")
                    print(f"  Fill color theme: '{getattr(cell.fill.start_color, 'theme', 'N/A')}'")
                    print(f"  Fill color tint:  '{getattr(cell.fill.start_color, 'tint', 'N/A')}'")
                    
                    # Comparar con una celda que SÍ sabemos que es festivo (ej. 01/01)
                    print(f"\n--- Comparación con celda Enero día 1 (festivo real) ---")
                    # Buscar bloque Enero
                    nc2 = 0
                    for r2 in range(1, ws.max_row + 1):
                        v2 = str(ws.cell(row=r2, column=1).value or '').strip()
                        if v2.upper() == "NOMBRE":
                            nc2 += 1
                            if nc2 == 1:  # Enero
                                for c2 in range(2, 33):
                                    if str(ws.cell(row=r2, column=c2).value).strip() == "1":
                                        for er2 in range(r2+1, r2+15):
                                            e2 = str(ws.cell(row=er2, column=1).value or '').strip()
                                            if e2.lower() == "eusebio":
                                                cell2 = ws.cell(row=er2, column=c2)
                                                print(f"  Valor:      '{cell2.value}'")
                                                print(f"  Fill type:  '{getattr(cell2.fill, 'fill_type', 'N/A')}'")
                                                print(f"  Fill color index: '{getattr(cell2.fill.start_color, 'index', 'N/A')}'")
                                                print(f"  Fill color rgb:   '{getattr(cell2.fill.start_color, 'rgb', 'N/A')}'")
                                                print(f"  Fill color type:  '{getattr(cell2.fill.start_color, 'type', 'N/A')}'")
                                                print(f"  Fill color theme: '{getattr(cell2.fill.start_color, 'theme', 'N/A')}'")
                                                break
                                        break
                    break
            break

input("\nPulsa Enter para cerrar...")
